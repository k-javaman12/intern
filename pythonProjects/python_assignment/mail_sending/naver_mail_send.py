from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl


def setup_webdriver():
    """Set up Chrome WebDriver with necessary options."""
    chrome_options = Options()
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    chrome_options.add_argument(f"user-agent={user_agent}")
    chrome_options.add_experimental_option("detach", True)
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
    return webdriver.Chrome(options=chrome_options)


def call_email(driver):
    """Function to send email using the given WebDriver instance."""
    wb = openpyxl.load_workbook('../email_list.xlsx')
    sheet = wb.active

    # Navigate to the URL
    driver.get("https://mail.naver.com/v2/new")

    # Click the button
    driver.find_element(By.XPATH, '//*[@id="root"]/div/nav/div/div[1]/div[2]/a[1]').click()

    for row in range(2, sheet.max_row + 1):
        # Assuming column 2 contains the email addresses
        email_address = sheet.cell(row=row, column=2).value

        # Enter the email address
        email_input = driver.find_element(By.XPATH, '//*[@id="recipient_input_element"]')
        email_input.clear()
        email_input.send_keys(email_address)

        # Write to the subject input
        subject_input = driver.find_element(By.XPATH, '//*[@id="subject_title"]')
        subject_input.clear()
        subject_input.send_keys("테스트")

        # Write to the body input
        body_input = driver.find_element(By.XPATH, '/html/body/div/div[1]')
        body_input.clear()
        body_input.send_keys("테스트 내용")

        # Click the send button (assuming it's a send button)
        driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div[1]/div/button[1]').click()


def main():
    """Main function to execute the script."""
    driver = setup_webdriver()
    call_email(driver)


if __name__ == "__main__":
    main()
