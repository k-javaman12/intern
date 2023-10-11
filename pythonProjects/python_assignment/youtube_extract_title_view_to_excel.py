import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Function to setup the Chrome WebDriver with the necessary options
def setup_webdriver():
    # Create an instance of Chrome Options
    chrome_options = Options()
    # Set a fake user-agent value to mimic a real browser request
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    chrome_options.add_argument(f"user-agent={user_agent}")

    # Ensure that browser remains open after the script finishes
    chrome_options.add_experimental_option("detach", True)

    # Disable the infobar that says "Chrome is being controlled by automated software"
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])

   #  Return the WebDriver instance with the specified options
    return webdriver.Chrome(options=chrome_options)


# Function to extract video titles, views, and links from the YouTube page
def get_videos_data(driver, url):
    # Direct the WebDriver to open the specified URL
    driver.get(url)

    # Define a constant for wait time between scrolling
    WAIT_IN_SECONDS = 5
    # Get the current height of the page
    last_height = driver.execute_script("return document.documentElement.scrollHeight")

    # Continuously scroll until the end of the page is reached
    while True:
        # Scroll to the bottom of the current page view
        driver.execute_script("window.scrollTo(0, arguments[0]);", last_height)
        # Wait a bit for more content to load
        time.sleep(WAIT_IN_SECONDS)

        # Get the new height of the page after scrolling
        new_height = driver.execute_script("return document.documentElement.scrollHeight")
        # If the new height is the same as the last height, we've reached the end
        if new_height == last_height:
            break
        last_height = new_height

    # Fetch video views, titles, and links
    views = driver.find_elements(By.XPATH, '//div[@id="metadata-line"]/span[1]')
    titles = driver.find_elements(By.ID, "video-title")
    links = driver.find_elements(By.ID, "video-title-link")

    # Use list comprehension to create a list of dictionaries for each video
    # In summary, zip() is a convenient way to combine data from multiple iterables when you want to process them together element-wise
    return [{'title': t.text, 'views': v.text, 'link': l.get_attribute('href')} for t, v, l in
            zip(titles, views, links)]


# Function to save the extracted data to an Excel file with specified column widths
def save_to_excel(data):
    # Convert the list of dictionaries to a pandas DataFrame
    df = pd.DataFrame(data)

    # Define the output filename
    filename = "youtube_videos.xlsx"

    # Save the DataFrame to an Excel file
    df.to_excel(filename, index=False, engine='openpyxl')

    # Open the Excel file for editing
    workbook = load_workbook(filename)
    worksheet = workbook.active

    # Set custom column widths (converted to Excel's width unit)
    col_widths = [600, 150, 300]
    for i, width in enumerate(col_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width / 6

    # Save the adjusted Excel file
    workbook.save(filename)


# Main function to execute the script
def main():
    # Set up the WebDriver
    driver = setup_webdriver()

    # Extract video data from the specified YouTube URL
    videos_data = get_videos_data(driver, "https://www.youtube.com/@ggassembly/videos")

    # Save the data to an Excel file
    save_to_excel(videos_data)


# Ensure that the script is only executed when run directly (not imported)
if __name__ == "__main__":
    main()
